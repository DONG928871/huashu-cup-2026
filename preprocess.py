# -*- coding: utf-8 -*-
"""
A题 数据预处理脚本
功能：读取附件Excel → 数据梳理 → 质量检查 → 预处理 → 可视化 → 输出
"""

import os, sys, warnings, json, io
warnings.filterwarnings('ignore')

# Fix Windows console encoding
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except:
        pass

# ============================================================
# 0. 路径与兼容读取
# ============================================================
XLSX_PATH = r"C:\Users\31670\OneDrive\Desktop\华数杯数学建模比赛\A题附件.xlsx"
OUT_DIR   = r"C:\Users\31670\OneDrive\Desktop\华数杯数学建模比赛\预处理输出"
os.makedirs(OUT_DIR, exist_ok=True)

def read_xlsx_robust(path):
    """
    兼容读取 .xlsx：优先 openpyxl，回退 zipfile+XML 解析
    返回: {sheet_name: [{col_letter: value}, ...]}
    """
    # --- 方法1: 尝试 openpyxl ---
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        result = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows_data = []
            for row in ws.iter_rows(min_row=1, values_only=False):
                row_dict = {}
                for cell in row:
                    if cell.value is not None:
                        row_dict[cell.column_letter] = cell.value
                if row_dict:
                    rows_data.append(row_dict)
            result[sn] = rows_data
        wb.close()
        print("[INFO] 使用 openpyxl 读取成功")
        return result
    except ImportError:
        print("[WARN] openpyxl 未安装，回退到 zipfile+XML 解析")
    except Exception as e:
        print(f"[WARN] openpyxl 读取失败 ({e})，回退到 zipfile+XML 解析")

    # --- 方法2: zipfile + XML 解析 ---
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(path, 'r') as z:
        # 解析共享字符串表
        ss_path = 'xl/sharedStrings.xml'
        strings = []
        if ss_path in z.namelist():
            with z.open(ss_path) as f:
                tree = ET.parse(f)
                ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in tree.findall('.//s:si', ns):
                    texts = [t.text or '' for t in si.findall('.//s:t', ns)]
                    strings.append(''.join(texts))

        # 解析工作表名称
        with z.open('xl/workbook.xml') as f:
            tree = ET.parse(f)
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            sheet_names = [s.get('name') for s in tree.findall('.//s:sheet', ns)]

        result = {}
        for idx, sname in enumerate(sheet_names, 1):
            sheet_file = f'xl/worksheets/sheet{idx}.xml'
            if sheet_file not in z.namelist():
                continue
            with z.open(sheet_file) as f:
                tree = ET.parse(f)
                ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                rows_data = []
                for row in tree.findall('.//s:row', ns):
                    row_dict = {}
                    for c in row.findall('s:c', ns):
                        ref = c.get('r')
                        col_letter = ''.join(ch for ch in ref if ch.isalpha())
                        v_el = c.find('s:v', ns)
                        t = c.get('t', '')
                        val = v_el.text if v_el is not None else ''
                        if t == 's' and val and val.isdigit():
                            idx_s = int(val)
                            if idx_s < len(strings):
                                val = strings[idx_s]
                        elif val:
                            try:
                                val = float(val) if '.' in val or 'e' in val.lower() else int(val)
                            except ValueError:
                                pass
                        row_dict[col_letter] = val
                    if row_dict:
                        rows_data.append(row_dict)
                result[sname] = rows_data
        print("[INFO] 使用 zipfile+XML 解析成功")
        return result

# ============================================================
# 1. 读取数据
# ============================================================
print("=" * 70)
print("阶段一：数据读取与结构梳理")
print("=" * 70)
raw_data = read_xlsx_robust(XLSX_PATH)
print(f"工作表数量: {len(raw_data)}")
for sn, rows in raw_data.items():
    print(f"  [{sn}] 行数: {len(rows)}, 首行键: {list(rows[0].keys()) if rows else '空'}")

# ============================================================
# 2. 数据结构化：拆分每个sheet的两组数据
# ============================================================
print("\n" + "=" * 70)
print("阶段二：数据结构化")
print("=" * 70)

# 根据读取结果分析结构
# 每个sheet: Row1=合并表头(问题1 A-C, 问题2 D-F), Row2=子表头(X,Y,Z ×2), Row3+=数据
structured = {}  # {sheet_name: {scenario: DataFrame-like dict}}

for sn, rows in raw_data.items():
    if not rows:
        continue
    # 跳过表头行（检查第一行是否为字符串标签）
    header_row1 = rows[0]

    # 识别两组数据: 组A (cols A,B,C) 和 组B (cols D,E,F)
    data_A = []  # 场景一
    data_B = []  # 场景二

    for row in rows:
        # 跳过表头行（值包含非数字的行）
        has_string = any(isinstance(v, str) for v in row.values())
        if has_string:
            continue

        # 提取数值
        row_A = {}
        row_B = {}
        for col in ['A', 'B', 'C']:
            if col in row and isinstance(row[col], (int, float)):
                row_A[col] = row[col]
        for col in ['D', 'E', 'F']:
            if col in row and isinstance(row[col], (int, float)):
                row_B[col] = row[col]

        if len(row_A) == 3:
            data_A.append(row_A)
        if len(row_B) == 3:
            data_B.append(row_B)

    structured[f"{sn}_场景A"] = data_A
    structured[f"{sn}_场景B"] = data_B
    print(f"  {sn}_场景A: {len(data_A)} 个粒子")
    print(f"  {sn}_场景B: {len(data_B)} 个粒子")

# 为方便建模，将数据重命名
final_datasets = {}
for key, data in structured.items():
    # 转换为 {X: [...], Y: [...], Z: [...]} 格式
    if not data:
        continue
    xs = [row.get('A', row.get('D', 0)) for row in data]
    ys = [row.get('B', row.get('E', 0)) for row in data]
    zs = [row.get('C', row.get('F', 0)) for row in data]
    final_datasets[key] = {'X': xs, 'Y': ys, 'Z': zs, 'N': len(data)}

print(f"\n  共生成 {len(final_datasets)} 个结构化数据集")

# ============================================================
# 3. 数据质量检查
# ============================================================
print("\n" + "=" * 70)
print("阶段三：数据质量检查")
print("=" * 70)

quality_report = {}
for name, ds in final_datasets.items():
    n = ds['N']
    xs, ys, zs = ds['X'], ds['Y'], ds['Z']

    # 缺失值检查
    nan_x = sum(1 for x in xs if x is None or (isinstance(x, float) and str(x) == 'nan'))
    nan_y = sum(1 for y in ys if y is None or (isinstance(y, float) and str(y) == 'nan'))
    nan_z = sum(1 for z in zs if z is None or (isinstance(z, float) and str(z) == 'nan'))

    # 重复值检查
    points = list(zip(xs, ys, zs))
    unique_points = len(set(points))
    duplicates = len(points) - unique_points

    # 统计量
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    z_min, z_max = min(zs), max(zs)
    x_mean = sum(xs) / n
    y_mean = sum(ys) / n
    z_mean = sum(zs) / n

    # IQR 异常值检测
    def iqr_bounds(vals):
        sorted_v = sorted(vals)
        q1 = sorted_v[n // 4]
        q3 = sorted_v[3 * n // 4]
        iqr = q3 - q1
        return q1 - 1.5 * iqr, q3 + 1.5 * iqr

    x_low, x_high = iqr_bounds(xs)
    y_low, y_high = iqr_bounds(ys)
    z_low, z_high = iqr_bounds(zs)

    x_outliers = sum(1 for x in xs if x < x_low or x > x_high)
    y_outliers = sum(1 for y in ys if y < y_low or y > y_high)
    z_outliers = sum(1 for z in zs if z < z_low or z > z_high)

    quality_report[name] = {
        'N': n,
        'missing_X': nan_x, 'missing_Y': nan_y, 'missing_Z': nan_z,
        'duplicates': duplicates,
        'X_range': (x_min, x_max), 'Y_range': (y_min, y_max), 'Z_range': (z_min, z_max),
        'X_mean': x_mean, 'Y_mean': y_mean, 'Z_mean': z_mean,
        'X_outliers': x_outliers, 'Y_outliers': y_outliers, 'Z_outliers': z_outliers
    }

    print(f"\n  [{name}]")
    print(f"    粒子数: {n}")
    print(f"    缺失值: X={nan_x}, Y={nan_y}, Z={nan_z}")
    print(f"    完全重复点: {duplicates}")
    print(f"    X 范围: [{x_min:.2f}, {x_max:.2f}], 均值: {x_mean:.2f}")
    print(f"    Y 范围: [{y_min:.2f}, {y_max:.2f}], 均值: {y_mean:.2f}")
    print(f"    Z 范围: [{z_min:.2f}, {z_max:.2f}], 均值: {z_mean:.2f}")
    print(f"    IQR异常值: X={x_outliers}, Y={y_outliers}, Z={z_outliers}")

# ============================================================
# 4. 预处理判断与执行
# ============================================================
print("\n" + "=" * 70)
print("阶段四：预处理执行")
print("=" * 70)

# 预处理决策矩阵
preprocess_log = []

for name, ds in final_datasets.items():
    qr = quality_report[name]
    log_entry = {'dataset': name, 'actions': []}

    # 4.1 缺失值检查
    total_missing = qr['missing_X'] + qr['missing_Y'] + qr['missing_Z']
    if total_missing > 0:
        log_entry['actions'].append(f"缺失值填充(总计{total_missing}个)")
        # 对于空间坐标，缺失意味着该行无效 → 删除整行（保守策略）
        print(f"  [{name}] 检测到 {total_missing} 个缺失值，将删除对应行")
    else:
        print(f"  [{name}] [OK] 无缺失值")
        log_entry['actions'].append("无需缺失值处理")

    # 4.2 异常值检查（基于IQR）
    total_outliers = qr['X_outliers'] + qr['Y_outliers'] + qr['Z_outliers']
    if total_outliers > 0:
        # 对于本题场景：坐标在[-5000,5000]范围内的点都是合理的RVE内部点
        # IQR的"异常值"可能是边界附近的合法点 → 保留，仅标记
        print(f"  [{name}] 检测到 IQR 统计异常值 {total_outliers} 个（坐标在RVE范围内，保留不删除）")
        log_entry['actions'].append(f"IQR异常值{total_outliers}个(保留-属RVE边界合法点)")
    else:
        print(f"  [{name}] [OK] 无IQR统计异常值")
        log_entry['actions'].append("无需异常值处理")

    # 4.3 重复值检查
    if qr['duplicates'] > 0:
        print(f"  [{name}] 检测到 {qr['duplicates']} 个完全重复点，将去重")
        log_entry['actions'].append(f"删除{qr['duplicates']}个重复点")
    else:
        print(f"  [{name}] [OK] 无重复点")
        log_entry['actions'].append("无需去重处理")

    # 4.4 归一化判断
    # 对于图连通性分析（基于距离），不需要归一化
    # 距离计算本身是尺度不变的
    print(f"  [{name}] → 距离型图模型，无需归一化/标准化")
    log_entry['actions'].append("无需归一化(距离型图模型)")

    # 4.5 编码判断
    # 坐标是连续数值，无分类变量，无需编码
    print(f"  [{name}] → 全连续数值变量，无需编码")
    log_entry['actions'].append("无需编码(全连续数值)")

    preprocess_log.append(log_entry)

# ============================================================
# 5. 实际预处理操作
# ============================================================
print("\n" + "=" * 70)
print("阶段五：执行处理操作")
print("=" * 70)

processed_data = {}
for name, ds in final_datasets.items():
    xs, ys, zs = list(ds['X']), list(ds['Y']), list(ds['Z'])
    n_orig = ds['N']

    # 5.1 删除含缺失值的行
    clean_x, clean_y, clean_z = [], [], []
    for x, y, z in zip(xs, ys, zs):
        valid = True
        for v in [x, y, z]:
            if v is None:
                valid = False
            elif isinstance(v, float) and (str(v) == 'nan' or v != v):
                valid = False
        if valid:
            clean_x.append(x)
            clean_y.append(y)
            clean_z.append(z)
    n_after_nan = len(clean_x)

    # 5.2 去重
    seen = set()
    dedup_x, dedup_y, dedup_z = [], [], []
    for x, y, z in zip(clean_x, clean_y, clean_z):
        key = (round(x, 8), round(y, 8), round(z, 8))  # 浮点容差去重
        if key not in seen:
            seen.add(key)
            dedup_x.append(x)
            dedup_y.append(y)
            dedup_z.append(z)
    n_after_dedup = len(dedup_x)

    processed_data[name] = {
        'X': dedup_x, 'Y': dedup_y, 'Z': dedup_z,
        'N_original': n_orig,
        'N_clean': n_after_nan,
        'N_final': n_after_dedup
    }

    print(f"  [{name}]: {n_orig} → (去缺失) → {n_after_nan} → (去重) → {n_after_dedup}")

# ============================================================
# 6. 输出处理后的数据
# ============================================================
print("\n" + "=" * 70)
print("阶段六：输出处理后的数据")
print("=" * 70)

for name, ds in processed_data.items():
    n = ds['N_final']
    xs, ys, zs = ds['X'], ds['Y'], ds['Z']

    print(f"\n{'-'*60}")
    print(f"  [{name}]  处理后: {n} 个粒子")
    print(f"{'-'*60}")
    print(f"  {'序号':>4s}  {'X':>16s}  {'Y':>16s}  {'Z':>16s}")
    print(f"  {'-'*60}")

    # 前10行
    show_n = min(10, n)
    for i in range(show_n):
        print(f"  {i+1:4d}  {xs[i]:16.6f}  {ys[i]:16.6f}  {zs[i]:16.6f}")

    if n > 15:
        print(f"  {'...':>4s}  {'...':>16s}  {'...':>16s}  {'...':>16s}")

    # 后5行
    for i in range(max(0, n-5), n):
        print(f"  {i+1:4d}  {xs[i]:16.6f}  {ys[i]:16.6f}  {zs[i]:16.6f}")

# ============================================================
# 7. 保存为CSV/MAT格式
# ============================================================
print("\n" + "=" * 70)
print("阶段七：保存数据文件")
print("=" * 70)

# CSV格式
for name, ds in processed_data.items():
    safe_name = name.replace(' ', '_')
    csv_path = os.path.join(OUT_DIR, f"{safe_name}.csv")
    with open(csv_path, 'w', encoding='utf-8-sig') as f:
        f.write("X,Y,Z\n")
        for x, y, z in zip(ds['X'], ds['Y'], ds['Z']):
            f.write(f"{x},{y},{z}\n")
    print(f"  [OK] CSV: {csv_path} ({ds['N_final']} 行)")

# JSON格式（便于程序读取）
json_path = os.path.join(OUT_DIR, "all_datasets.json")
json_data = {}
for name, ds in processed_data.items():
    json_data[name] = {
        'X': ds['X'], 'Y': ds['Y'], 'Z': ds['Z'],
        'N': ds['N_final'],
        'N_original': ds['N_original'],
        'description': '微构体导电填料三维坐标'
    }
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)
print(f"  [OK] JSON: {json_path}")

# 预处理报告
report_path = os.path.join(OUT_DIR, "预处理报告.json")
with open(report_path, 'w', encoding='utf-8') as f:
    json.dump({
        'quality_report': {k: {kk: (str(vv) if isinstance(vv, tuple) else vv) for kk, vv in v.items()} for k, v in quality_report.items()},
        'preprocess_actions': preprocess_log,
        'final_summary': {name: {'N_original': ds['N_original'], 'N_final': ds['N_final']} for name, ds in processed_data.items()}
    }, f, ensure_ascii=False, indent=2)
print(f"  [OK] 报告: {report_path}")

# ============================================================
# 8. 数据完整性校验
# ============================================================
print("\n" + "=" * 70)
print("阶段八：数据完整性校验")
print("=" * 70)

all_checks_pass = True
for name, ds in processed_data.items():
    n = ds['N_final']
    xs, ys, zs = ds['X'], ds['Y'], ds['Z']

    checks = []
    # 校验1：行列一致性
    c1 = len(xs) == len(ys) == len(zs) == n
    checks.append(('行列一致性', c1))

    # 校验2：无非有限值
    import math
    c2 = all(math.isfinite(v) for v in xs + ys + zs)
    checks.append(('有限数值', c2))

    # 校验3：坐标范围合理性（RVE边界 [-5000, 5000]）
    c3 = all(-5000.1 <= v <= 5000.1 for v in xs + ys + zs)
    checks.append(('RVE边界', c3))

    # 校验4：粒子数 > 0
    c4 = n > 0
    checks.append(('非空数据集', c4))

    # 校验5：无重复
    points_set = set((round(x, 8), round(y, 8), round(z, 8)) for x, y, z in zip(xs, ys, zs))
    c5 = len(points_set) == n
    checks.append(('无重复点', c5))

    all_ok = all(c[1] for c in checks)
    all_checks_pass = all_checks_pass and all_ok
    status = "[OK] 通过" if all_ok else "[FAIL] 失败"
    print(f"\n  [{name}] 完整性校验: {status}")
    for check_name, result in checks:
        print(f"    {'[OK]' if result else '[FAIL]'} {check_name}")

print(f"\n  总体校验: {'[OK] 全部通过' if all_checks_pass else '[FAIL] 存在未通过的校验项'}")

# ============================================================
# 9. 可视化
# ============================================================
print("\n" + "=" * 70)
print("阶段九：生成数据预处理可视化")
print("=" * 70)

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from mpl_toolkits.mplot3d import Axes3D
    import numpy as np

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    # 选择代表性数据集进行可视化（问题1场景A和场景B）
    vis_datasets = [k for k in processed_data.keys() if '问题1' in k or 'Sheet1' in k]
    if not vis_datasets:
        vis_datasets = list(processed_data.keys())[:2]

    n_vis = len(vis_datasets)
    fig = plt.figure(figsize=(8 * n_vis, 14))

    for vi, name in enumerate(vis_datasets):
        ds = processed_data[name]
        xs = np.array(ds['X'])
        ys = np.array(ds['Y'])
        zs = np.array(ds['Z'])

        # ---- 子图1: 预处理前（原始数据 = 处理后数据，因为数据本身就干净）----
        # 为了展示预处理对比，我们人为记录统计分布
        ax1 = fig.add_subplot(2, n_vis, vi + 1, projection='3d')
        ax1.scatter(xs, ys, zs, c='#2a78d6', s=30, alpha=0.7, edgecolors='none', label='粒子坐标')
        ax1.set_title(f'{name}\n(预处理后 — 共 {len(xs)} 个粒子)', fontsize=11, fontweight='bold', pad=12)
        ax1.set_xlabel('X', fontsize=9)
        ax1.set_ylabel('Y', fontsize=9)
        ax1.set_zlabel('Z', fontsize=9)

        # 绘制 RVE 边界框
        def draw_rve_box(ax, size=5000):
            corners = np.array([[-size, -size, -size], [size, -size, -size],
                               [size, size, -size], [-size, size, -size],
                               [-size, -size, size], [size, -size, size],
                               [size, size, size], [-size, size, size]])
            edges = [(0,1),(1,2),(2,3),(3,0),(4,5),(5,6),(6,7),(7,4),(0,4),(1,5),(2,6),(3,7)]
            for e in edges:
                ax.plot3D(*zip(corners[e[0]], corners[e[1]]), color='#898781', linewidth=0.5, linestyle='--', alpha=0.5)

        draw_rve_box(ax1)
        ax1.set_xlim(-5500, 5500)
        ax1.set_ylim(-5500, 5500)
        ax1.set_zlim(-5500, 5500)
        ax1.view_init(elev=20, azim=45)

        # ---- 子图2: 各维度分布直方图 ----
        ax2 = fig.add_subplot(2, n_vis, vi + 1 + n_vis)

        # 这里比较"预处理前后"——由于本数据无需预处理，展示分布一致性
        colors = {'X': '#2a78d6', 'Y': '#eb6834', 'Z': '#1baf7a'}
        for coord, vals, color in [('X', xs, colors['X']), ('Y', ys, colors['Y']), ('Z', zs, colors['Z'])]:
            ax2.hist(vals, bins=25, alpha=0.45, color=color, edgecolor='white', linewidth=0.3,
                    label=f'{coord} (σ={np.std(vals):.0f})')

        ax2.set_title(f'{name} — 坐标分布直方图', fontsize=11, fontweight='bold')
        ax2.set_xlabel('坐标值', fontsize=9)
        ax2.set_ylabel('频数', fontsize=9)
        ax2.legend(fontsize=8, loc='upper right')
        ax2.axvline(x=-5000, color='#898781', linewidth=0.8, linestyle='--', alpha=0.5)
        ax2.axvline(x=5000, color='#898781', linewidth=0.8, linestyle='--', alpha=0.5)
        ax2.grid(True, alpha=0.3, linewidth=0.5)

    plt.suptitle('A题 微构体导电填料数据 — 预处理结果可视化', fontsize=14, fontweight='bold', y=0.98)
    plt.tight_layout(rect=[0, 0, 1, 0.96])

    img_path = os.path.join(OUT_DIR, '数据预处理可视化.png')
    plt.savefig(img_path, dpi=200, bbox_inches='tight', facecolor='#fcfcfb')
    plt.close()
    print(f"  [OK] 可视化已保存: {img_path}")

except ImportError as e:
    print(f"  [WARN] matplotlib 不可用 ({e})，跳过可视化")
    print(f"  [INFO] 请手动运行: pip install matplotlib")

print("\n" + "=" * 70)
print("预处理完成！")
print(f"输出目录: {OUT_DIR}")
print("=" * 70)
